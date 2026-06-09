from pathlib import Path

import openpyxl

from config import Config
from models import VulnerabilityTask


class TaskLoader:
    def __init__(self, config: Config):
        self.config = config
        self.excel_path = config.root_dir / config.excel_path

    def load_tasks(self) -> list[VulnerabilityTask]:
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)

        # Load project metadata from summary sheet
        projects = self._load_projects(wb["汇总"])

        tasks = []
        for project_meta in projects:
            project_name = project_meta["project"]
            if project_name not in wb.sheetnames:
                continue
            ws = wb[project_name]
            project_tasks = self._load_project_tasks(ws, project_meta)
            tasks.extend(project_tasks)

        wb.close()
        return tasks

    def _load_projects(self, ws) -> list[dict]:
        projects = []
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(v).strip().lower() if v else "" for v in row]
                continue
            values = list(row)
            if not values or not values[0]:
                continue
            meta = {}
            for j, key in enumerate(header):
                if j < len(values):
                    meta[key] = values[j]
            # Ensure required fields
            if meta.get("project") and meta.get("github_url"):
                projects.append({
                    "project": str(meta["project"]),
                    "github_url": str(meta["github_url"]),
                    "owner": str(meta.get("owner", "")),
                    "repo": str(meta.get("repo", "")),
                })
        return projects

    def _load_project_tasks(self, ws, project_meta: dict) -> list[VulnerabilityTask]:
        tasks = []
        header = None
        data_start = 9  # Row 9 is header (0-indexed: 8)

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == data_start - 1:
                # Header row
                header = [str(v).strip().lower() if v else "" for v in row]
                continue
            if i < data_start:
                continue

            values = list(row)
            if not values or not values[0]:
                continue

            # Build record
            record = {}
            for j, key in enumerate(header):
                if j < len(values):
                    record[key] = values[j]

            task = self._build_task(project_meta, record)
            if task:
                tasks.append(task)

        return tasks

    def _build_task(self, project_meta: dict, record: dict) -> VulnerabilityTask | None:
        source = str(record.get("source", "")).strip()
        cve_id = record.get("cve-id")
        adv_id = record.get("adv-id")

        # Convert to string or None
        cve_id = str(cve_id).strip() if cve_id else None
        adv_id = str(adv_id).strip() if adv_id else None

        # Skip rows without any ID
        if not cve_id and not adv_id:
            return None

        # Canonical ID
        canonical_id = cve_id if cve_id else adv_id

        # Task key
        task_key = f"{project_meta['project']}:{source}:{canonical_id}"

        return VulnerabilityTask(
            project=project_meta["project"],
            github_url=project_meta["github_url"],
            owner=project_meta["owner"],
            repo=project_meta["repo"],
            source=source,
            cve_id=cve_id,
            adv_id=adv_id,
            canonical_id=canonical_id,
            task_key=task_key,
            publish_at=str(record.get("publish-at", "")) if record.get("publish-at") else None,
            cwe=str(record.get("cwe", "")) if record.get("cwe") else None,
        )
